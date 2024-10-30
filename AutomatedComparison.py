'''
Author: DengYH
Date: 2024-10-30 11:16:11
'''
import os
import re
from openpyxl import load_workbook

def find_and_write_matched_data(dir_path):
    os.chdir(dir_path)
    files = os.listdir()
    pattern = re.compile(r'\d*404\d*')  # 正则表达式，匹配学号数据(此处可根据个人情况进行修改)
    matched_data = []
    for file in files:
        match = pattern.search(file)
        if match:
            matched_data.append(match.group())
        else:
            matched_data.append("")
    matched_data.sort()
    return matched_data
    
def target_student():
    try:
        wb = load_workbook(filename='此处填入学号姓名文件路径.xlsx')
        sheet = wb['Sheet1']
        column_data = []
        for row in sheet.iter_rows(values_only=True):
            column_data.append(row[1])
        return column_data
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return []

def find_remaining_data(column_data, new_matched_data):
    remaining_data = []
    for item in column_data:
        if item is not None and str(item) not in new_matched_data:
            remaining_data.append(item)
    return remaining_data

if __name__ == "__main__":
    dir_path = "此处选择需要比对的文档路径"
    new_matched_data = find_and_write_matched_data(dir_path)
    column_data = target_student()
    remaining_data = find_remaining_data(column_data, new_matched_data)
    # print("已交作业学号：",new_matched_data)
    print(dir_path)
    print("未交学号：",remaining_data)
    print("总人数：",len(column_data)," ； 已交人数：",len(new_matched_data)," ； 缺交人数：",len(remaining_data))
    # print("总名单：",column_data)